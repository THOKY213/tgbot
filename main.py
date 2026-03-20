"""
Бот для напоминания о задолженностях
Запуск: pip install python-telegram-bot openpyxl xlrd odfpy && python debt_bot.py
"""

import sqlite3
import logging
import logging.handlers
import traceback
import io
import csv

import openpyxl

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler, CallbackQueryHandler,
    ConversationHandler, ContextTypes, filters
)
from telegram.constants import ParseMode

# ─── НАСТРОЙКИ ────────────────────────────────────────────────────────────────

BOT_TOKEN    = "8785317982:AAEZFtCiFN5jYJE4GHcQGuwZgtSVZoJngto"
DB           = "debts.db"
LOG_FILE     = "bot.log"   # путь к файлу логов, или "" чтобы не писать в файл
LOG_ADMIN_ID = 938917446        # ваш Telegram ID — сюда будут приходить уведомления об ошибках

# ─── СОСТОЯНИЯ ────────────────────────────────────────────────────────────────

ADD_NAME, ADD_PHONE, ADD_AMOUNT, ADD_DESC, ADD_DATE = range(5)
SEND_MSG  = 10
WAIT_FILE = 20

# ─── ЛОГИРОВАНИЕ ──────────────────────────────────────────────────────────────

def setup_logging():
    fmt = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    root = logging.getLogger()
    root.setLevel(logging.INFO)

    # Консоль
    ch = logging.StreamHandler()
    ch.setFormatter(fmt)
    root.addHandler(ch)

    # Файл с ротацией: макс 5 МБ, хранить 3 файла
    if LOG_FILE:
        fh = logging.handlers.RotatingFileHandler(
            LOG_FILE, maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8"
        )
        fh.setFormatter(fmt)
        root.addHandler(fh)

    # Убираем лишний шум от httpx
    logging.getLogger("httpx").setLevel(logging.WARNING)

setup_logging()
logger = logging.getLogger(__name__)


async def notify_admin_error(bot, text: str):
    """Отправляет уведомление об ошибке администратору в Telegram."""
    if not LOG_ADMIN_ID:
        return
    try:
        msg = "🚨 *Ошибка бота*\n\n" + text[:3000]
        await bot.send_message(chat_id=LOG_ADMIN_ID, text=msg, parse_mode=ParseMode.MARKDOWN)
    except Exception:
        pass


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    """Глобальный обработчик необработанных исключений."""
    err = context.error
    tb = "".join(traceback.format_exception(type(err), err, err.__traceback__))
    user_info = ""
    if isinstance(update, Update) and update.effective_user:
        u = update.effective_user
        user_info = f" | user={u.id} @{u.username or '—'}"
    logger.error(f"Unhandled exception{user_info}:\n{tb}")
    await notify_admin_error(context.bot, tb)
    if isinstance(update, Update) and update.effective_message:
        try:
            await update.effective_message.reply_text(
                "⚠️ Произошла ошибка. Попробуйте ещё раз или нажмите /start"
            )
        except Exception:
            pass

# ─── БАЗА ДАННЫХ ──────────────────────────────────────────────────────────────

def init_db():
    with sqlite3.connect(DB) as c:
        c.execute("""
            CREATE TABLE IF NOT EXISTS clients (
                id      INTEGER PRIMARY KEY AUTOINCREMENT,
                name    TEXT NOT NULL,
                phone   TEXT,
                amount  REAL NOT NULL,
                desc    TEXT,
                due     TEXT
            )
        """)
        cols = [row[1] for row in c.execute("PRAGMA table_info(clients)")]
        if "name" not in cols or "amount" not in cols:
            c.execute("DROP TABLE clients")
            c.execute("""
                CREATE TABLE clients (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL, phone TEXT,
                    amount REAL NOT NULL, desc TEXT, due TEXT
                )
            """)

def db(sql, params=()):
    with sqlite3.connect(DB) as c:
        c.row_factory = sqlite3.Row
        return c.execute(sql, params).fetchall()

def db1(sql, params=()):
    with sqlite3.connect(DB) as c:
        c.row_factory = sqlite3.Row
        return c.execute(sql, params).fetchone()

def dbx(sql, params=()):
    with sqlite3.connect(DB) as c:
        return c.execute(sql, params).lastrowid

# ─── КНОПКИ ───────────────────────────────────────────────────────────────────

def main_menu_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Добавить клиента", callback_data="add"),
         InlineKeyboardButton("📋 Список клиентов",  callback_data="list")],
        [InlineKeyboardButton("📢 Напоминания",       callback_data="remind"),
         InlineKeyboardButton("📥 Импорт файла",      callback_data="import")],
        [InlineKeyboardButton("🗑 Удалить клиента",   callback_data="del_menu")],
    ])

def back_kb():
    return InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Главное меню", callback_data="back_main")]])

def back_and_action_kb(label, cb):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(label, callback_data=cb)],
        [InlineKeyboardButton("◀️ Главное меню", callback_data="back_main")],
    ])

# ─── СЕССИЯ: защита от потери данных при перезапуске ─────────────────────────

def session_lost(update):
    """Отправляет сообщение если сессия диалога потерялась."""
    return update.message.reply_text(
        "⚠️ Сессия прервалась (бот перезапускался).\nНачните заново — нажмите /start",
        reply_markup=main_menu_kb()
    )

# ─── ГЛАВНОЕ МЕНЮ ─────────────────────────────────────────────────────────────

async def show_main(message, edit=False):
    text = "🤖 *Бот напоминания о долгах*\n\nВыберите действие:"
    if edit:
        await message.edit_text(text, reply_markup=main_menu_kb(), parse_mode=ParseMode.MARKDOWN)
    else:
        await message.reply_text(text, reply_markup=main_menu_kb(), parse_mode=ParseMode.MARKDOWN)

async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    logger.info(f"[start] user={u.id} @{u.username or '—'} name={u.full_name}")
    await show_main(update.effective_message)

async def back_handler(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    ctx.user_data.clear()
    await show_main(q.message, edit=True)

# ─── СПИСОК КЛИЕНТОВ ──────────────────────────────────────────────────────────

async def list_clients(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    rows = db("SELECT * FROM clients ORDER BY id")
    if not rows:
        await q.message.edit_text(
            "📋 Список клиентов пуст.\n\nДобавьте клиентов вручную или импортируйте из файла.",
            reply_markup=back_and_action_kb("➕ Добавить клиента", "add"),
            parse_mode=ParseMode.MARKDOWN
        )
        return

    total_sum = sum(r["amount"] for r in rows)
    text = f"📋 *Клиентов: {len(rows)} | Долгов на: {total_sum:,.2f} руб.*\n{'─'*32}\n\n"
    for r in rows:
        text += (
            f"🆔`{r['id']}` *{r['name']}*\n"
            f"   📞 {r['phone'] or '—'}  💰 {r['amount']:,.2f} руб.\n"
            f"   📋 {r['desc'] or '—'}  📅 {r['due'] or '—'}\n\n"
        )
    text += "_/del <ID> — удалить клиента_"
    await q.message.edit_text(text, reply_markup=back_kb(), parse_mode=ParseMode.MARKDOWN)

# ─── ДОБАВЛЕНИЕ КЛИЕНТА ───────────────────────────────────────────────────────

async def add_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    ctx.user_data.clear()
    await q.message.edit_text(
        "➕ *Добавление клиента* (1/5)\n\n✏️ Введите *ФИО* клиента:\n\n/cancel — отмена",
        reply_markup=None,
        parse_mode=ParseMode.MARKDOWN
    )
    return ADD_NAME

async def add_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["name"] = update.message.text.strip()
    await update.message.reply_text(
        "➕ *Добавление клиента* (2/5)\n\n📞 Введите *телефон*:\n\n/skip — пропустить  /cancel — отмена",
        parse_mode=ParseMode.MARKDOWN
    )
    return ADD_PHONE

async def add_phone(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if "name" not in ctx.user_data:
        await session_lost(update)
        return ConversationHandler.END
    txt = update.message.text.strip()
    ctx.user_data["phone"] = None if txt == "/skip" else txt
    await update.message.reply_text(
        "➕ *Добавление клиента* (3/5)\n\n💰 Введите *сумму долга* (руб.):\n\n/cancel — отмена",
        parse_mode=ParseMode.MARKDOWN
    )
    return ADD_AMOUNT

async def add_amount(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if "name" not in ctx.user_data:
        await session_lost(update)
        return ConversationHandler.END
    try:
        ctx.user_data["amount"] = float(update.message.text.strip().replace(",", ".").replace(" ", ""))
    except ValueError:
        await update.message.reply_text("❌ Введите число, например: *15000* или *1500.50*", parse_mode=ParseMode.MARKDOWN)
        return ADD_AMOUNT
    await update.message.reply_text(
        "➕ *Добавление клиента* (4/5)\n\n📋 Введите *описание* долга:\n\n/skip — пропустить  /cancel — отмена",
        parse_mode=ParseMode.MARKDOWN
    )
    return ADD_DESC

async def add_desc(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if "amount" not in ctx.user_data:
        await session_lost(update)
        return ConversationHandler.END
    txt = update.message.text.strip()
    ctx.user_data["desc"] = None if txt == "/skip" else txt
    await update.message.reply_text(
        "➕ *Добавление клиента* (5/5)\n\n📅 Введите *дату оплаты* (напр. 31.03.2025):\n\n/skip — пропустить  /cancel — отмена",
        parse_mode=ParseMode.MARKDOWN
    )
    return ADD_DATE

async def add_date(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if "name" not in ctx.user_data or "amount" not in ctx.user_data:
        await session_lost(update)
        return ConversationHandler.END
    txt = update.message.text.strip()
    due = None if txt == "/skip" else txt
    d = ctx.user_data
    cid = dbx(
        "INSERT INTO clients (name, phone, amount, desc, due) VALUES (?,?,?,?,?)",
        (d["name"], d.get("phone"), d["amount"], d.get("desc"), due)
    )
    logger.info(f"[add_client] id={cid} name={d['name']} amount={d['amount']} by user={update.effective_user.id}")
    ctx.user_data.clear()
    await update.message.reply_text(
        f"✅ *Клиент добавлен!*\n\n"
        f"🆔 ID: `{cid}`\n"
        f"👤 {d['name']}\n"
        f"📞 {d.get('phone') or '—'}\n"
        f"💰 {d['amount']:,.2f} руб.\n"
        f"📋 {d.get('desc') or '—'}\n"
        f"📅 {due or '—'}",
        reply_markup=back_and_action_kb("➕ Добавить ещё", "add"),
        parse_mode=ParseMode.MARKDOWN
    )
    return ConversationHandler.END

# ─── УДАЛЕНИЕ ─────────────────────────────────────────────────────────────────

async def del_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    rows = db("SELECT * FROM clients ORDER BY id")
    if not rows:
        await q.message.edit_text("📋 Нет клиентов для удаления.", reply_markup=back_kb())
        return
    text = "🗑 *Удаление клиента*\n\nВведите команду:\n`/del <ID>`\n\n"
    for r in rows:
        text += f"🆔`{r['id']}` {r['name']} — {r['amount']:,.0f} руб.\n"
    await q.message.edit_text(text, reply_markup=back_kb(), parse_mode=ParseMode.MARKDOWN)

async def cmd_del(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not ctx.args:
        await update.message.reply_text("Использование: `/del <ID>`", parse_mode=ParseMode.MARKDOWN)
        return
    try:
        cid = int(ctx.args[0])
    except ValueError:
        await update.message.reply_text("❌ ID должен быть числом")
        return
    client = db1("SELECT * FROM clients WHERE id=?", (cid,))
    if not client:
        await update.message.reply_text(f"❌ Клиент `{cid}` не найден.", parse_mode=ParseMode.MARKDOWN)
        return
    logger.info(f"[delete] client_id={cid} name={client['name']} by user={update.effective_user.id}")
    dbx("DELETE FROM clients WHERE id=?", (cid,))
    await update.message.reply_text(
        f"✅ *{client['name']}* удалён.",
        reply_markup=back_and_action_kb("🗑 Удалить ещё", "del_menu"),
        parse_mode=ParseMode.MARKDOWN
    )

# ─── РАССЫЛКА ─────────────────────────────────────────────────────────────────

async def remind_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    rows = db("SELECT * FROM clients ORDER BY id")
    if not rows:
        await q.message.edit_text("📋 Нет клиентов для рассылки.", reply_markup=back_kb())
        return
    kb = [[InlineKeyboardButton(f"{r['name']} — {r['amount']:,.0f} руб.", callback_data=f"r_{r['id']}")]
          for r in rows]
    kb.append([InlineKeyboardButton("📢 Всем сразу", callback_data="r_all")])
    kb.append([InlineKeyboardButton("◀️ Главное меню", callback_data="back_main")])
    await q.message.edit_text(
        "📢 *Напоминания*\n\nКому отправить?",
        reply_markup=InlineKeyboardMarkup(kb),
        parse_mode=ParseMode.MARKDOWN
    )

async def remind_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    target = q.data
    if target == "r_all":
        ctx.user_data["remind_ids"] = [r["id"] for r in db("SELECT id FROM clients")]
        hint = f"для *{len(ctx.user_data['remind_ids'])} клиентов*"
    else:
        cid = int(target.split("_")[1])
        client = db1("SELECT * FROM clients WHERE id=?", (cid,))
        ctx.user_data["remind_ids"] = [cid]
        hint = f"для *{client['name']}*"
    await q.message.edit_text(
        f"📢 *Текст напоминания* {hint}\n\n"
        f"Подстановки: `{{name}}` `{{amount}}` `{{due}}`\n\n"
        f"Введите текст:\n\n/cancel — отмена",
        parse_mode=ParseMode.MARKDOWN
    )
    return SEND_MSG

async def remind_send(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if "remind_ids" not in ctx.user_data:
        await session_lost(update)
        return ConversationHandler.END
    template = update.message.text.strip()
    ids = ctx.user_data.get("remind_ids", [])
    lines = []
    for cid in ids:
        r = db1("SELECT * FROM clients WHERE id=?", (cid,))
        if not r:
            continue
        msg = (template
               .replace("{name}", r["name"])
               .replace("{amount}", f"{r['amount']:,.2f} руб.")
               .replace("{due}", r["due"] or "—"))
        lines.append(f"📨 *{r['name']}* ({r['phone'] or 'нет тел.'}):\n`{msg}`")
    logger.info(f"[remind] sent={len(lines)} by user={update.effective_user.id}")
    result = f"✅ *Готово! Напоминаний: {len(lines)}*\n{'─'*30}\n\n"
    result += "\n\n".join(lines[:10])
    if len(lines) > 10:
        result += f"\n\n_...и ещё {len(lines) - 10}_"
    ctx.user_data.clear()
    await update.message.reply_text(
        result,
        reply_markup=back_and_action_kb("📢 Новая рассылка", "remind"),
        parse_mode=ParseMode.MARKDOWN
    )
    return ConversationHandler.END

# ─── ИМПОРТ ФАЙЛА ─────────────────────────────────────────────────────────────

ALIASES = {
    "name": [
        "имя", "фио", "ф.и.о", "ф.и.о.", "клиент", "заказчик", "должник",
        "контрагент", "наименование", "организация", "компания", "фирма",
        "плательщик", "получатель", "партнёр", "партнер", "покупатель",
        "арендатор", "физлицо", "юрлицо", "абонент", "пользователь",
        "name", "full name", "fullname", "client", "customer", "debtor",
        "company", "organization", "organisation", "contractor", "partner",
        "payer", "tenant", "account", "contact name", "borrower", "person",
    ],
    "phone": [
        "телефон", "тел", "тел.", "моб", "мобильный", "номер", "номер телефона",
        "контакт", "сотовый", "мобильный телефон", "телефон клиента", "связь",
        "whatsapp", "вотсап", "вайбер", "viber",
        "phone", "phone number", "tel", "mobile", "cell", "contact",
        "telephone", "cellphone", "number", "mob",
    ],
    "amount": [
        "сумма", "долг", "задолженность", "к оплате", "итого", "баланс",
        "оплата", "сумма долга", "сумма задолженности", "сумма к оплате",
        "остаток", "остаток долга", "долг клиента", "сумма платежа",
        "начислено", "задолжал", "недоплата", "недостача", "взыскать",
        "ущерб", "штраф", "пеня", "неустойка", "кредит", "займ", "заем",
        "amount", "debt", "balance", "sum", "total", "due", "owed",
        "outstanding", "payment", "invoice amount", "charge", "fee",
        "liability", "arrears", "overdue amount", "loan", "credit",
    ],
    "desc": [
        "описание", "назначение", "комментарий", "примечание", "услуга",
        "товар", "продукт", "за что", "основание", "договор", "счёт",
        "счет", "номер договора", "№ договора", "позиция", "статья",
        "предмет", "вид долга", "тип", "категория", "детали", "инфо",
        "description", "note", "comment", "service", "product", "goods",
        "details", "memo", "reason", "purpose", "invoice", "contract",
        "item", "category", "type", "info", "remarks", "reference",
    ],
    "due": [
        "дата", "срок", "до", "оплатить до", "дата оплаты", "срок оплаты",
        "срок погашения", "дата погашения", "крайний срок", "крайняя дата",
        "дата платежа", "платёж до", "платеж до", "дедлайн", "до какого",
        "дата задолженности", "срок долга", "план оплаты", "ожидается",
        "due", "due date", "deadline", "date", "payment date", "pay by",
        "expiry", "expiration", "maturity", "maturity date", "until",
        "expected", "target date", "schedule", "pay date", "end date",
    ],
}

def detect_col(header):
    h = str(header).strip().lower()
    for field, aliases in ALIASES.items():
        if any(h == a or h.startswith(a) for a in aliases):
            return field
    return None

def parse_rows_from_table(rows_iter):
    rows = list(rows_iter)
    if not rows:
        return None, [], "Файл пуст."
    col_map = {}
    for idx, h in enumerate(rows[0]):
        if h is None:
            continue
        field = detect_col(str(h))
        if field and field not in col_map:
            col_map[field] = idx
    if "name" in col_map and "amount" in col_map:
        detected = {f: rows[0][i] for f, i in col_map.items()}
        detected_msg = ("✅ *Определил колонки:*\n"
                        + "\n".join(f"  • {f} → «{v}»" for f, v in detected.items()) + "\n\n")
        data_rows = [(i + 2, row) for i, row in enumerate(rows[1:])]
    else:
        col_map = {"name": 0, "phone": 1, "amount": 2, "desc": 3, "due": 4}
        detected_msg = "⚠️ Заголовки не распознаны — читаю по порядку: ФИО, Телефон, Сумма, Описание, Дата.\n\n"
        data_rows = [(i + 1, row) for i, row in enumerate(rows)]
    return col_map, data_rows, detected_msg

def import_rows(col_map, data_rows):
    imported, errors = 0, []
    def get(row, field):
        idx = col_map.get(field)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if v is not None else None
    for line_num, row in data_rows:
        if not any(v for v in row if v is not None):
            continue
        name   = get(row, "name") or ""
        phone  = get(row, "phone")
        amount = get(row, "amount")
        desc   = get(row, "desc")
        due    = get(row, "due")
        if not name:
            errors.append(f"Строка {line_num}: пустое имя")
            continue
        try:
            amount = float(str(amount).replace(",", ".").replace(" ", "").replace("\xa0", ""))
        except (TypeError, ValueError, AttributeError):
            errors.append(f"Строка {line_num} ({name}): некорректная сумма «{amount}»")
            continue
        dbx("INSERT INTO clients (name, phone, amount, desc, due) VALUES (?,?,?,?,?)",
            (name, phone or None, amount, desc, due))
        imported += 1
    return imported, errors

def load_xlsx(data):
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    return [[cell.value for cell in row] for row in wb.active.iter_rows()]

def load_xlsm(data):
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, keep_vba=False)
    return [[cell.value for cell in row] for row in wb.active.iter_rows()]

def load_xls(data):
    import xlrd
    wb = xlrd.open_workbook(file_contents=data)
    ws = wb.sheet_by_index(0)
    return [ws.row_values(i) for i in range(ws.nrows)]

def load_ods(data):
    from odf.opendocument import load as ods_load
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
    doc = ods_load(io.BytesIO(data))
    sheet = doc.spreadsheet.getElementsByType(Table)[0]
    rows = []
    for tr in sheet.getElementsByType(TableRow):
        row = []
        for tc in tr.getElementsByType(TableCell):
            ps = tc.getElementsByType(P)
            row.append(str(ps[0]) if ps else None)
        rows.append(row)
    return rows

def load_csv(data):
    for enc in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = data.decode("utf-8", errors="replace")
    dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    return list(csv.reader(io.StringIO(text), dialect))

SUPPORTED_FORMATS = {
    ".xlsx": load_xlsx,
    ".xlsm": load_xlsm,
    ".xls":  load_xls,
    ".ods":  load_ods,
    ".csv":  load_csv,
}

async def import_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    await q.message.edit_text(
        "📥 *Импорт клиентов*\n\n"
        "*Форматы:* `.xlsx` `.xlsm` `.xls` `.ods` `.csv`\n\n"
        "*Обязательные колонки:*\n"
        "• ФИО / Имя / Клиент / Должник / name / client...\n"
        "• Сумма / Долг / Задолженность / amount / debt...\n\n"
        "*Необязательные:*\n"
        "• Телефон / phone  •  Описание / description  •  Дата / due date\n\n"
        "📎 Отправьте файл:\n\n/cancel — отмена",
        reply_markup=None,
        parse_mode=ParseMode.MARKDOWN
    )
    return WAIT_FILE

async def import_file(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not doc:
        await update.message.reply_text("❌ Отправьте файл.")
        return WAIT_FILE
    fname = doc.file_name.lower()
    ext = next((e for e in SUPPORTED_FORMATS if fname.endswith(e)), None)
    if not ext:
        await update.message.reply_text(
            f"❌ Формат не поддерживается.\nПоддерживаются: {', '.join(SUPPORTED_FORMATS)}"
        )
        return WAIT_FILE
    data = bytes(await (await doc.get_file()).download_as_bytearray())
    try:
        raw_rows = SUPPORTED_FORMATS[ext](data)
    except ImportError as e:
        lib = str(e).split("'")[1] if "'" in str(e) else str(e)
        await update.message.reply_text(f"❌ Нужна библиотека: `pip install {lib}`", parse_mode=ParseMode.MARKDOWN)
        return WAIT_FILE
    except Exception as e:
        await update.message.reply_text(f"❌ Не удалось прочитать файл: {e}")
        return WAIT_FILE

    col_map, data_rows, detected_msg = parse_rows_from_table(raw_rows)
    if col_map is None or "name" not in col_map or "amount" not in col_map:
        await update.message.reply_text(
            "❌ Не найдены обязательные колонки.\n\n"
            "Назовите их: *ФИО/Клиент/Должник* и *Сумма/Долг/Задолженность*",
            parse_mode=ParseMode.MARKDOWN
        )
        return WAIT_FILE

    imported, errors = import_rows(col_map, data_rows)
    logger.info(f"[import] file={fname} imported={imported} errors={len(errors)} by user={update.effective_user.id}")
    msg = detected_msg + f"📊 Загружено: *{imported}*\n❌ Ошибок: *{len(errors)}*"
    if errors:
        msg += "\n\n*Ошибки:*\n" + "\n".join(f"  • {e}" for e in errors[:10])
    await update.message.reply_text(
        msg,
        reply_markup=back_and_action_kb("📥 Загрузить ещё", "import"),
        parse_mode=ParseMode.MARKDOWN
    )
    return ConversationHandler.END

# ─── ОТМЕНА ───────────────────────────────────────────────────────────────────

async def cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    await update.message.reply_text("❌ Отменено.", reply_markup=main_menu_kb())
    return ConversationHandler.END

# ─── РОУТЕР КНОПОК ────────────────────────────────────────────────────────────

async def button_router(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    data = q.data
    if data == "list":
        await list_clients(update, ctx)
    elif data == "del_menu":
        await del_menu(update, ctx)
    elif data == "remind":
        await remind_menu(update, ctx)
    elif data.startswith("back_"):
        ctx.user_data.clear()
        await show_main(q.message, edit=True)

# ─── ЗАПУСК ───────────────────────────────────────────────────────────────────

def main():
    init_db()
    logger.info("🤖 Бот запускается...")
    app = Application.builder().token(BOT_TOKEN).build()

    add_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(add_start, pattern="^add$")],
        per_message=False,
        states={
            ADD_NAME:   [MessageHandler(filters.TEXT & ~filters.COMMAND, add_name)],
            ADD_PHONE:  [MessageHandler(filters.TEXT & ~filters.COMMAND, add_phone),
                         CommandHandler("skip", add_phone)],
            ADD_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_amount)],
            ADD_DESC:   [MessageHandler(filters.TEXT & ~filters.COMMAND, add_desc),
                         CommandHandler("skip", add_desc)],
            ADD_DATE:   [MessageHandler(filters.TEXT & ~filters.COMMAND, add_date),
                         CommandHandler("skip", add_date)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    remind_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(remind_menu, pattern="^remind$")],
        per_message=False,
        states={
            SEND_MSG: [MessageHandler(filters.TEXT & ~filters.COMMAND, remind_send)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    import_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(import_start, pattern="^import$")],
        per_message=False,
        states={
            WAIT_FILE: [MessageHandler(filters.Document.ALL, import_file)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_error_handler(error_handler)
    app.add_handler(add_conv)
    app.add_handler(remind_conv)
    app.add_handler(import_conv)
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("del", cmd_del))
    app.add_handler(CommandHandler("cancel", cancel))
    app.add_handler(CallbackQueryHandler(remind_pick, pattern="^r_"))
    app.add_handler(CallbackQueryHandler(button_router))

    logger.info("✅ Бот запущен!")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
